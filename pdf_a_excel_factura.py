import os
import re
import io
import shutil
import argparse
from pathlib import Path
from datetime import datetime, timedelta
from typing import Optional, Tuple
import time

import openpyxl
import pdfplumber
from openpyxl.styles import Alignment

# OCR fallback (sin Poppler)
try:
    import fitz  # pymupdf
    import pytesseract
    from PIL import Image
    OCR_OK = True
except Exception:
    OCR_OK = False


# ================= CONFIG =================
FILAS_INICIO, FILAS_FIN = 9, 21
PROVINCIA_FIJA = "Mendoza"

IMPRIMIR_AUTOMATICO = True
COPIAS_IMPRESION = 1
ESPERA_IMPRESION_SEG = 0.5

MESES_ES = {
    1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril",
    5: "Mayo", 6: "Junio", 7: "Julio", 8: "Agosto",
    9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre"
}


# ================= UTILIDADES =================
def normalizar_texto(text: str) -> str:
    text = text or ""
    text = text.replace("\u00a0", " ")
    text = text.replace("–", "-").replace("—", "-").replace("−", "-")
    text = text.replace("º", "°")
    text = "\n".join([ln.strip() for ln in text.splitlines() if ln.strip()])
    text = re.sub(r"[ \t]+", " ", text)
    return text


def strip_leading_zeros(num_str: str) -> str:
    s = re.sub(r"[^\d]", "", num_str or "")
    s = s.lstrip("0")
    return s if s else "0"


def only_digits(s: str) -> str:
    return re.sub(r"\D", "", s or "")


def format_cuit(val: str) -> str:
    d = only_digits(val)
    if len(d) == 11:
        return f"{d[:2]}-{d[2:10]}-{d[10:]}"
    return val or ""


def monto_ar_to_float(s: str) -> float:
    """Formato AR típico: 1.234.567,89"""
    s = (s or "").strip()
    if not s:
        return 0.0
    s = s.replace(".", "").replace(",", ".")
    s = re.sub(r"[^\d\.\-]", "", s)
    if s in ("", "-", ".", "-."):
        return 0.0
    try:
        return float(s)
    except Exception:
        return 0.0


def monto_to_float_any(s: str) -> float:
    """
    Soporta:
      - AR: 1.234.567,89
      - US/INT: 1,234,567.89
      - Simple: 125200 / 125200,00 / 125200.00
    """
    s = (s or "").strip()
    if not s:
        return 0.0

    s = re.sub(r"[^\d\.,\-]", "", s)

    # US/INT: coma miles + punto decimales
    if re.fullmatch(r"-?\d{1,3}(?:,\d{3})*\.\d{2}", s):
        try:
            return float(s.replace(",", ""))
        except Exception:
            return 0.0

    # AR: punto miles + coma decimales
    if re.fullmatch(r"-?\d{1,3}(?:\.\d{3})*,\d{2}", s):
        return monto_ar_to_float(s)

    # Solo coma o solo punto
    if re.fullmatch(r"-?\d+,\d{2}", s):
        try:
            return float(s.replace(",", "."))
        except Exception:
            return 0.0
    if re.fullmatch(r"-?\d+\.\d{2}", s):
        try:
            return float(s)
        except Exception:
            return 0.0

    # Entero (o mezcla rara): quitamos separadores y parseamos
    try:
        return float(s.replace(",", ""))
    except Exception:
        return 0.0


def sanitize_filename(name: str) -> str:
    name = re.sub(r'[\\/:*?"<>|]+', " ", name or "")
    name = re.sub(r"\s+", " ", name).strip()
    return name if name else "SIN_NOMBRE"


def unique_path(path: Path) -> Path:
    if not path.exists():
        return path
    i = 2
    while True:
        new = path.with_stem(f"{path.stem} ({i})")
        if not new.exists():
            return new
        i += 1


def parse_date_flexible(s: str) -> datetime:
    return datetime.strptime(s.strip(), "%d/%m/%Y")


def parse_date_dash(s: str) -> datetime:
    return datetime.strptime(s.strip(), "%d-%m-%Y")


def month_start(dt: Optional[datetime]) -> Optional[datetime]:
    if not dt:
        return None
    return dt.replace(day=1, hour=0, minute=0, second=0, microsecond=0)


def prev_month_start(dt: Optional[datetime]) -> Optional[datetime]:
    if not dt:
        return None
    first = dt.replace(day=1)
    prev_last = first - timedelta(days=1)
    return prev_last.replace(day=1, hour=0, minute=0, second=0, microsecond=0)


def wrap_by_words(text: str, max_len: int = 38) -> str:
    text = re.sub(r"\s+", " ", (text or "")).strip()
    if len(text) <= max_len:
        return text
    words = text.split(" ")
    lines, cur = [], ""
    for w in words:
        if not cur:
            cur = w
        elif len(cur) + 1 + len(w) <= max_len:
            cur += " " + w
        else:
            lines.append(cur)
            cur = w
    if cur:
        lines.append(cur)
    return "\n".join(lines)


def set_proveedor_en_A8(ws, razon_social: str):
    wrapped = wrap_by_words(razon_social, max_len=38)
    ws["A8"].value = wrapped
    ws["A8"].alignment = Alignment(wrap_text=True, vertical="top")
    lineas = max(1, wrapped.count("\n") + 1)
    ws.row_dimensions[8].height = 15 * lineas


def save_debug_text(error_dir: Path, pdf_path: Path, text: str):
    try:
        error_dir.mkdir(parents=True, exist_ok=True)
        dbg = error_dir / f"DEBUG_{pdf_path.stem}.txt"
        dbg.write_text(text or "", encoding="utf-8", errors="ignore")
    except Exception:
        pass


def save_parse_debug(error_dir: Path, pdf_obj: Path, parser_name: str, campos: dict, text: str):
    try:
        error_dir.mkdir(parents=True, exist_ok=True)
        p = error_dir / f"PARSE_{pdf_obj.stem}_{parser_name}.txt"
        keys = ["nro", "razon", "cuit", "iibb", "domicilio", "localidad", "provincia", "fecha", "periodo", "total"]
        contenido = [f"PARSER={parser_name}", "=== CAMPOS ==="]
        for k in keys:
            contenido.append(f"{k}: {campos.get(k)}")
        contenido.append("\n=== TEXTO ===\n")
        contenido.append(text or "")
        p.write_text("\n".join(contenido), encoding="utf-8", errors="ignore")
    except Exception:
        pass


# ================= VALIDACIÓN (CRÍTICOS) =================
def validar_campos_criticos(campos: dict) -> Tuple[bool, str]:
    faltantes = []
    if not (campos.get("razon") or "").strip():
        faltantes.append("Razón social")
    if not (campos.get("nro") or "").strip():
        faltantes.append("Número de factura")
    if not campos.get("fecha"):
        faltantes.append("Fecha de emisión")
    total = campos.get("total")
    if total is None or float(total) <= 0:
        faltantes.append("Total")
    if faltantes:
        return False, ", ".join(faltantes)
    return True, ""


# ================= DOMICILIO -> (domicilio, localidad, provincia) =================
def split_domicilio_localidad_provincia(texto_dom: str) -> Tuple[str, str, str]:
    s = re.sub(r"\s+", " ", (texto_dom or "")).strip()

    # cortar basura (NO debe quedar CUIT dentro del domicilio)
    s = re.split(r"\bCUIT\b\s*:", s, flags=re.IGNORECASE)[0].strip()
    s = re.split(r"\bIngresos\s+Brutos\b\s*:", s, flags=re.IGNORECASE)[0].strip()

    domicilio = s
    localidad = ""
    provincia = PROVINCIA_FIJA

    if " - " in s:
        left, right = s.split(" - ", 1)
        domicilio = left.strip()
        right = right.strip()

        if "," in right:
            loc, prov = right.split(",", 1)
            localidad = loc.strip()
            provincia = prov.strip() or PROVINCIA_FIJA
        else:
            if re.search(r"\bMendoza\b", right, re.IGNORECASE):
                localidad = re.sub(r"\bMendoza\b", "", right, flags=re.IGNORECASE).strip(" -,")
    else:
        parts = [p.strip() for p in s.split(",") if p.strip()]
        if len(parts) >= 3:
            domicilio = parts[0]
            localidad = parts[1]
            provincia = parts[2] or PROVINCIA_FIJA
        elif len(parts) == 2:
            domicilio = parts[0]
            if parts[1].lower() != "mendoza":
                localidad = parts[1]

    domicilio = domicilio.strip(" -,")
    if not provincia:
        provincia = PROVINCIA_FIJA

    return domicilio, localidad, provincia


# ================= TOTAL EN LETRAS (A40) =================
_UNIDADES = ["", "uno", "dos", "tres", "cuatro", "cinco", "seis", "siete", "ocho", "nueve"]
_DIEZ_A_19 = ["diez", "once", "doce", "trece", "catorce", "quince", "dieciséis", "diecisiete", "dieciocho", "diecinueve"]
_DECENAS = ["", "", "veinte", "treinta", "cuarenta", "cincuenta", "sesenta", "setenta", "ochenta", "noventa"]
_CIENTOS = ["", "ciento", "doscientos", "trescientos", "cuatrocientos", "quinientos",
            "seiscientos", "setecientos", "ochocientos", "novecientos"]


def _hasta_99(n: int) -> str:
    if n < 10:
        return _UNIDADES[n]
    if 10 <= n < 20:
        return _DIEZ_A_19[n - 10]
    if 20 <= n < 30:
        if n == 20:
            return "veinte"
        return "veinti" + _UNIDADES[n - 20]
    d = n // 10
    u = n % 10
    if u == 0:
        return _DECENAS[d]
    return f"{_DECENAS[d]} y {_UNIDADES[u]}"


def _hasta_999(n: int) -> str:
    if n == 0:
        return ""
    if n == 100:
        return "cien"
    c = n // 100
    r = n % 100
    if c == 0:
        return _hasta_99(r)
    if r == 0:
        return _CIENTOS[c] if c != 1 else "cien"
    return f"{_CIENTOS[c]} {_hasta_99(r)}".strip()


def numero_a_letras_es(n: int) -> str:
    if n == 0:
        return "cero"
    partes = []
    millones = n // 1_000_000
    n = n % 1_000_000
    miles = n // 1000
    resto = n % 1000

    if millones:
        if millones == 1:
            partes.append("un millón")
        else:
            partes.append(f"{numero_a_letras_es(millones)} millones")

    if miles:
        if miles == 1:
            partes.append("mil")
        else:
            partes.append(f"{_hasta_999(miles)} mil".strip())

    if resto:
        partes.append(_hasta_999(resto).strip())

    return " ".join([p for p in partes if p]).strip()


def total_pesos_a_letras(total: float) -> str:
    total = float(total or 0.0)
    if total < 0:
        total = abs(total)

    pesos = int(total)
    centavos = int(round((total - pesos) * 100))
    if centavos == 100:
        pesos += 1
        centavos = 0

    texto_pesos = numero_a_letras_es(pesos)
    texto_pesos = re.sub(r"\buno\b$", "un", texto_pesos)
    texto_pesos = re.sub(r"veintiuno\b", "veintiún", texto_pesos)
    texto_pesos = re.sub(r" y uno\b", " y un", texto_pesos)

    moneda = "peso" if pesos == 1 else "pesos"
    return f"{texto_pesos} {moneda} con {centavos:02d}/100".upper()


# ================= IMPRESIÓN =================
def imprimir_excel_windows(ruta_excel: str, copies: int = 1) -> bool:
    try:
        import win32com.client  # type: ignore
        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        wb = excel.Workbooks.Open(os.path.abspath(ruta_excel))
        wb.PrintOut(Copies=int(copies))
        time.sleep(ESPERA_IMPRESION_SEG)
        wb.Close(SaveChanges=False)
        excel.Quit()
        return True
    except Exception:
        pass

    try:
        os.startfile(os.path.abspath(ruta_excel), "print")
        return True
    except Exception:
        return False


# ================= LECTURA PDF + OCR =================
def read_pdf_text_first_page(pdf_path: str) -> str:
    try:
        with pdfplumber.open(pdf_path) as pdf:
            txt = pdf.pages[0].extract_text() or ""
        txt = normalizar_texto(txt)
        if txt.strip():
            return txt
    except Exception:
        pass

    if not OCR_OK:
        return ""

    try:
        doc = fitz.open(pdf_path)
        page = doc.load_page(0)
        zoom = 3
        pix = page.get_pixmap(matrix=fitz.Matrix(zoom, zoom))
        img = Image.open(io.BytesIO(pix.tobytes("png")))

        for lang in ("spa+eng", "spa", "eng"):
            try:
                ocr = pytesseract.image_to_string(img, lang=lang) or ""
                ocr = normalizar_texto(ocr)
                if ocr.strip():
                    return ocr
            except Exception:
                continue
        return ""
    except Exception:
        return ""


# ================= PARSERS =================
def _find_money_amounts_any(lines):
    amts = []
    for ln in lines:
        for m in re.findall(r"\b\d{1,3}(?:\.\d{3})*,\d{2}\b", ln):  # AR
            amts.append(monto_to_float_any(m))
        for m in re.findall(r"\b\d{1,3}(?:,\d{3})*\.\d{2}\b", ln):  # US/INT
            amts.append(monto_to_float_any(m))
        for m in re.findall(r"\b\d{2,}\,\d{2}\b", ln):  # 12345,67
            amts.append(monto_to_float_any(m))
        for m in re.findall(r"\b\d{2,}\.\d{2}\b", ln):  # 12345.67
            amts.append(monto_to_float_any(m))
        for m in re.findall(r"\b\d{2,}\b", ln):  # enteros "grandes"
            if len(m) >= 5:
                amts.append(monto_to_float_any(m))
    return [a for a in amts if a and a > 0]


def _is_bad_singleton(ln: str) -> bool:
    s = (ln or "").strip().upper()
    if not s:
        return True
    if s in ("B", "A", "C", "FACTURA", "B FACTURA", "ORIGINAL", "DUPLICADO", "TRIPLICADO"):
        return True
    if re.fullmatch(r"COD\.\s*\d+", s):
        return True
    if re.fullmatch(r"[A-Z]\s*FACTURA", s):
        return True
    return False


def _compact_ocr_token_text(text: str) -> str:
    """
    Compacta separaciones OCR raras dentro de palabras:
    ej: "C FACTU RA" -> "CFACTURA", "Emisi ón" -> "Emisión" (aprox. por regex flexible posterior).
    """
    s = text or ""
    prev = None
    while prev != s:
        prev = s
        s = re.sub(r"\b([A-Za-zÁÉÍÓÚÜÑ])\s+(?=[A-Za-zÁÉÍÓÚÜÑ]\b)", r"\1", s)
    return s


def parse_formato_ama(text: str):
    """
    A.M.A - Asociación Mendocina de Anestesiología
    Ej:
      B FACTURA
      N°.: 0021 - 00005279
      Fecha de Emisión: 04/03/2026
      CUIT.:30-64519120-6
      ING. BRUTOS.:4367
      DON BOSCO 474 – CIUDAD DE MENDOZA
      Período 3 / 2026
      Neto Exento: $6.854.083,80 ***** ... $6.854.083,80
    """
    if not text:
        return None

    lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
    raw = "\n".join(lines)
    up = re.sub(r"\s+", " ", raw).upper()

    if "ASOCIACIÓN MENDOCINA DE ANESTESIOLOG" not in up and "ASOCIACION MENDOCINA DE ANESTESIOLOG" not in up and "A.M.A" not in up:
        return None
    if "FACTURA" not in up:
        return None

    m = re.search(r"N[°º]?\.\s*:\s*(\d{4})\s*-\s*(\d{8})", raw, re.IGNORECASE)
    if not m:
        m = re.search(r"\b(\d{4})\s*-\s*(\d{8})\b", raw)
    if not m:
        return None
    nro = f"{strip_leading_zeros(m.group(1))}-{strip_leading_zeros(m.group(2))}"

    fecha = None
    fm = re.search(r"Fecha\s+de\s+Emisi[oó]n\s*:\s*(\d{1,2}/\d{1,2}/\d{4})", raw, re.IGNORECASE)
    if fm:
        try:
            fecha = parse_date_flexible(fm.group(1))
        except Exception:
            fecha = None

    razon = "Asociación Mendocina de Anestesiología"
    for ln in lines[:25]:
        if re.search(r"Asociaci[oó]n\s+Mendocina\s+de\s+Anestesiolog", ln, re.IGNORECASE):
            razon = re.sub(r"\s+", " ", ln).strip()
            break

    cuit = ""
    cm = re.search(r"CUIT\.\s*:\s*([0-9]{2}\s*[-]?\s*[0-9]{8}\s*[-]?\s*[0-9])", raw, re.IGNORECASE)
    if cm:
        cuit = format_cuit(cm.group(1))

    iibb = ""
    im = re.search(r"ING\.\s*BRUTOS\.\s*:\s*([0-9\.\-]+)", raw, re.IGNORECASE)
    if im:
        iibb = im.group(1).strip()

    domicilio = ""
    localidad = ""
    provincia = PROVINCIA_FIJA

    # domicilio línea con "DON BOSCO 474"
    for ln in lines:
        if re.search(r"\bDON\s+BOSCO\b", ln, re.IGNORECASE):
            domicilio = re.sub(r"\s+", " ", ln).strip()
            break

    # si viene "Mendoza - Mendoza"
    for ln in lines:
        if re.search(r"\bMENDOZA\s*-\s*MENDOZA\b", ln, re.IGNORECASE):
            localidad = "Mendoza"
            provincia = "Mendoza"
            break

    # Período "3 / 2026"
    periodo = None
    pm = re.search(r"Per[ií]odo\s+(\d{1,2})\s*/\s*(\d{4})", raw, re.IGNORECASE)
    if pm:
        try:
            mm = int(pm.group(1))
            yy = int(pm.group(2))
            periodo = datetime(yy, mm, 1)
        except Exception:
            periodo = None

    # TOTAL: tomar el último $ de la línea "Neto Exento: $... ***** Total I.V.A ... $..."
    total = None
    for ln in lines:
        if re.search(r"Neto\s+Exento\s*:", ln, re.IGNORECASE) and "$" in ln:
            nums = re.findall(r"\$\s*([0-9\.\,]+)", ln)
            if nums:
                total = monto_to_float_any(nums[-1])
                break

    # fallback: si el PDF trae una línea suelta con "$6.854.083,80"
    if total is None or total <= 0:
        amts = _find_money_amounts_any(lines)
        total = max(amts) if amts else None

    if total is None or total <= 0:
        return None

    return {
        "nro": nro,
        "razon": razon,
        "cuit": cuit,
        "iibb": iibb,
        "domicilio": domicilio,
        "localidad": localidad,
        "provincia": provincia or PROVINCIA_FIJA,
        "fecha": fecha,
        "periodo": periodo,
        "total": float(total)
    }


def parse_formato_mk_g3(text: str):
    """
    MK & G3 S.A - Factura B N°: 0051-00001011
    Saca el TOTAL desde la línea que arranca con '$' y tiene varios '$' (tabla NETO...TOTAL).
    """
    if not text:
        return None

    lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
    raw = "\n".join(lines)
    up = re.sub(r"\s+", " ", raw).upper()

    if "FACTURA B" not in up:
        return None
    if not (("MK & G3" in up) or ("MK & G3 S.A" in up) or ("MK" in up and "G3" in up)):
        return None

    m = re.search(r"Factura\s*B\s*N[°º]?\s*:\s*(\d{4})-(\d{8})", raw, re.IGNORECASE)
    if not m:
        return None
    nro = f"{strip_leading_zeros(m.group(1))}-{strip_leading_zeros(m.group(2))}"

    fecha = None
    fm = re.search(r"\bFecha\s*:\s*(\d{1,2}/\d{1,2}/\d{4})\b", raw, re.IGNORECASE)
    if fm:
        try:
            fecha = parse_date_flexible(fm.group(1))
        except Exception:
            fecha = None

    razon = "MK & G3 S.A"
    for ln in lines[:25]:
        if re.search(r"\bMK\s*&\s*G3\b", ln, re.IGNORECASE):
            razon = re.sub(r"\s+", " ", ln).strip()
            break

    cuit = ""
    cm = re.search(r"\bCUIT\s*:\s*([0-9]{11})\b", raw, re.IGNORECASE)
    if cm:
        cuit = format_cuit(cm.group(1))

    iibb = ""
    im = re.search(r"\bING\.\s*BRUTOS\s*:\s*([0-9\.\-]+)\b", raw, re.IGNORECASE)
    if im:
        iibb = im.group(1).strip()

    domicilio = ""
    for i, ln in enumerate(lines[:35]):
        if re.search(r"\bMK\s*&\s*G3\b", ln, re.IGNORECASE):
            if i + 1 < len(lines):
                cand = re.sub(r"\s+", " ", lines[i + 1]).strip()
                if not re.search(r"\bCUIT\b|\bING\.\s*BRUTOS\b|FACTURA|FECHA", cand, re.IGNORECASE):
                    domicilio = cand
            break

    total = None
    for ln in lines:
        if ln.strip().startswith("$") and ln.count("$") >= 2:
            nums = re.findall(r"\$\s*([0-9\.\,]+)", ln)
            if nums:
                total = monto_to_float_any(nums[-1])
                break

    if total is None or total <= 0:
        tm = re.search(r"\bTOTAL\b.*?\$?\s*([0-9\.\,]+)", raw, re.IGNORECASE)
        if tm:
            total = monto_to_float_any(tm.group(1))

    if total is None or total <= 0:
        amts = _find_money_amounts_any(lines)
        total = max(amts) if amts else None

    if total is None or total <= 0:
        return None

    return {
        "nro": nro,
        "razon": razon,
        "cuit": cuit,
        "iibb": iibb,
        "domicilio": domicilio,
        "localidad": "",
        "provincia": PROVINCIA_FIJA,
        "fecha": fecha,
        "periodo": None,
        "total": float(total)
    }


def parse_formato_mendosalud(text: str):
    """
    MENDOSALUD - Factura B (formato específico).
    Detección con anclas combinadas para evitar colisiones.
    """
    if not text:
        return None

    lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
    raw = "\n".join(lines)
    up = re.sub(r"\s+", " ", raw).upper()

    if "MENDOSALUD" not in up:
        return None
    if "FACTURA B" not in up:
        return None
    if not re.search(r"ING\.?\s*BRUTOS", up, re.IGNORECASE):
        return None
    if not re.search(r"NRO\.?\s*ESTAB", up, re.IGNORECASE):
        return None
    if "CAE VTO" not in up:
        return None

    m = re.search(r"Factura\s*B\s*(?:N[°º]|NRO)\s*:\s*(\d+)\s*-\s*(\d+)", raw, re.IGNORECASE)
    if not m:
        return None
    nro = f"{strip_leading_zeros(m.group(1))}-{strip_leading_zeros(m.group(2))}"

    razon = "MENDOSALUD"

    cuit = ""
    cm = re.search(
        r"MENDOSALUD[\s\S]{0,120}?C\.?\s*U\.?\s*I\.?\s*T\.?\s*:\s*([0-9][0-9\-\.\s]{9,16}[0-9])",
        raw,
        re.IGNORECASE
    )
    if cm:
        cuit = format_cuit(cm.group(1))

    iibb = ""
    im = re.search(r"ING\.?\s*BRUTOS\.?\s*:\s*([0-9\.\-]+)", raw, re.IGNORECASE)
    if im:
        iibb = im.group(1).strip()

    fecha = None
    fm = re.search(r"\bFecha\s*:\s*(\d{1,2}/\d{1,2}/\d{4})", raw, re.IGNORECASE)
    if fm:
        try:
            fecha = parse_date_flexible(fm.group(1))
        except Exception:
            fecha = None

    domicilio = ""
    localidad = ""
    provincia = PROVINCIA_FIJA

    for ln in lines[:35]:
        if re.search(r"ING\.?\s*BRUTOS", ln, re.IGNORECASE):
            base = re.split(r"ING\.?\s*BRUTOS", ln, flags=re.IGNORECASE)[0].strip(" -,")
            base = re.sub(r"\(\d{4,5}\)", "", base).strip(" -,")
            partes = [p.strip(" .,-") for p in base.split(",") if p.strip(" .,-")]
            if partes:
                domicilio = partes[0]
                if len(partes) >= 3:
                    localidad = f"{partes[1]}, {partes[2]}"
                elif len(partes) >= 2:
                    localidad = partes[1]
                if len(partes) >= 4:
                    prov_raw = partes[3].upper()
                    if prov_raw in ("MZA", "MENDOZA"):
                        provincia = "Mendoza"
                    else:
                        provincia = partes[3]
            break

    total = None
    for i, ln in enumerate(lines):
        if re.search(r"NETO\s+EXENTO", ln, re.IGNORECASE) and re.search(r"TOTAL", ln, re.IGNORECASE):
            for j in range(i + 1, min(i + 4, len(lines))):
                if "$" in lines[j]:
                    nums = re.findall(r"\$\s*([0-9\.\,]+)", lines[j])
                    if nums:
                        total = monto_to_float_any(nums[-1])
                        break
            if total is not None and total > 0:
                break

    if total is None or total <= 0:
        tm = re.search(r"\bTOTAL\b[^\n]*\n[^\n]*\$\s*([0-9\.\,]+)\s*$", raw, re.IGNORECASE)
        if tm:
            total = monto_to_float_any(tm.group(1))

    if total is None or total <= 0:
        amts = _find_money_amounts_any(lines)
        total = max(amts) if amts else None

    if total is None or total <= 0:
        return None

    return {
        "nro": nro,
        "razon": razon,
        "cuit": cuit,
        "iibb": iibb,
        "domicilio": domicilio,
        "localidad": localidad,
        "provincia": provincia or PROVINCIA_FIJA,
        "fecha": fecha,
        "periodo": None,
        "total": float(total)
    }


def parse_formato_hospital_universitario(text: str):
    """
    Hospital Universitario - formato "C FACTURA 0234-00002977"
    (Solo primera página)
    """
    if not text:
        return None

    lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
    raw = "\n".join(lines)
    t = re.sub(r"\s+", " ", raw).upper()

    if "HOSPITAL UNIVERSITARIO" not in t or "FACTURA" not in t:
        return None

    m = re.search(r"\b(\d{4})\s*[- ]\s*(\d{8})\b", t)
    if not m:
        return None
    nro = f"{strip_leading_zeros(m.group(1))}-{strip_leading_zeros(m.group(2))}"

    fecha = None
    fm = re.search(r"\bFECHA\s*:\s*(\d{1,2}/\d{1,2}/\d{4})\b", raw, re.IGNORECASE)
    if fm:
        try:
            fecha = parse_date_flexible(fm.group(1))
        except Exception:
            fecha = None

    razon = "Hospital Universitario"

    cuit = ""
    cm = re.search(r"\bCUIT\s*:\s*([0-9]{11})\b", raw, re.IGNORECASE)
    if cm:
        cuit = format_cuit(cm.group(1))

    iibb = ""
    im = re.search(r"Ingresos\s*brutos\s*:\s*([0-9\.\-]+)", raw, re.IGNORECASE)
    if im:
        iibb = im.group(1).strip()

    domicilio = ""
    localidad = ""
    provincia = PROVINCIA_FIJA

    for ln in lines[:20]:
        if re.search(r"PASO\s+DE\s+LOS\s+ANDES", ln, re.IGNORECASE):
            domicilio = re.sub(r"\s+", " ", ln).strip()
            break

    for ln in lines[:25]:
        mloc = re.search(r"\b(\d{3,5})\s*,\s*([A-ZÁÉÍÓÚÜÑ ]+)\s*,\s*([A-ZÁÉÍÓÚÜÑ ]+)\b", ln.upper())
        if mloc:
            nro_calle = mloc.group(1).strip()
            if domicilio and not re.search(r"\b\d{2,5}\b", domicilio):
                domicilio = f"{domicilio} {nro_calle}".strip()
            elif not domicilio:
                domicilio = nro_calle
            localidad = mloc.group(2).strip().title()
            provincia = mloc.group(3).strip().title()
            break

    total = None
    tm = re.search(r"Total\s+General.*?\$?\s*([0-9\.\,]+)", raw, re.IGNORECASE)
    if tm:
        total = monto_to_float_any(tm.group(1))

    if total is None or total <= 0:
        tm2 = re.search(r"\bSERVICIOS\s+PRESTADOS\s*\$?\s*([0-9\.\,]+)", raw, re.IGNORECASE)
        if tm2:
            total = monto_to_float_any(tm2.group(1))

    if total is None or total <= 0:
        tm3 = re.search(r"Subtotal.*?\$?\s*([0-9\.\,]+)", raw, re.IGNORECASE)
        if tm3:
            total = monto_to_float_any(tm3.group(1))

    if total is None or total <= 0:
        amts = _find_money_amounts_any(lines)
        total = max(amts) if amts else None

    if total is None or total <= 0:
        return None

    return {
        "nro": nro,
        "razon": razon,
        "cuit": cuit,
        "iibb": iibb,
        "domicilio": domicilio,
        "localidad": localidad,
        "provincia": provincia or PROVINCIA_FIJA,
        "fecha": fecha,
        "periodo": None,
        "total": float(total)
    }


def parse_formato_nutrihome(text: str):
    """
    Nutri Home S.A. / Nutrihome (Nutrición domiciliaria)
    (Solo primera página)
    """
    if not text:
        return None

    lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
    raw = "\n".join(lines)
    t = re.sub(r"\s+", " ", raw).upper()

    if not (("NUTRIHOME" in t) or ("NUTRI HOME" in t) or ("NUTRICION DOMICILIARIA" in t)):
        return None
    if "FACTURA" not in t:
        return None

    m = re.search(r"\bFACTURA\s+(\d{5})\s*[- ]\s*(\d{8})\b", t, re.IGNORECASE)
    if not m:
        m = re.search(r"\b(\d{5})\s*[- ]\s*(\d{8})\b", t)
    if not m:
        return None
    nro = f"{strip_leading_zeros(m.group(1))}-{strip_leading_zeros(m.group(2))}"

    razon = ""
    for ln in lines[:25]:
        if re.search(r"\bNUTRI\s*HOME\b", ln, re.IGNORECASE):
            razon = re.sub(r"\s+", " ", ln).strip()
            break
    if not razon:
        razon = "NUTRI HOME S.A."

    cuit = ""
    cm = re.search(r"C\.?\s*U\.?\s*I\.?\s*T\.?\s*:\s*([0-9\-\.\s]{11,14})", raw, re.IGNORECASE)
    if cm:
        cuit = format_cuit(cm.group(1))

    iibb = ""
    im = re.search(r"Ing\.?\s*Brutos.*?:\s*([0-9\-\.\s]+)", raw, re.IGNORECASE)
    if im:
        iibb = re.sub(r"\s+", "", im.group(1)).strip()

    fecha = None
    fm = re.search(r"Fecha\s+de\s+Facturac\w*\s*:\s*(\d{1,2}/\d{1,2}/\d{4})", raw, re.IGNORECASE)
    if fm:
        try:
            fecha = parse_date_flexible(fm.group(1))
        except Exception:
            fecha = None

    domicilio = ""
    localidad = ""
    provincia = PROVINCIA_FIJA

    for ln in lines[:40]:
        if re.search(r"\bAV\.\b|\bAVENIDA\b|\bCABILD", ln, re.IGNORECASE):
            ln_clean = re.sub(r"\s+", " ", ln).strip()
            ln_clean = re.split(r"\bCODIGO\b|\bIMP\.\s*INTERNOS\b", ln_clean, flags=re.IGNORECASE)[0].strip()
            ln_clean = re.sub(r"\s*-\s*\d{4}.*$", "", ln_clean).strip()
            domicilio = ln_clean
            break

    if re.search(r"\bCABA\b|\bCAPITAL\s+FEDERAL\b", raw, re.IGNORECASE):
        localidad = "CABA"
        provincia = "CABA"

    total = None
    tm = re.search(r"\b(IMPORTE\s+TOTAL|TOTAL)\b\s*[:$]?\s*([0-9\.,]+)", raw, re.IGNORECASE)
    if tm:
        total = monto_to_float_any(tm.group(2))

    if total is None or total <= 0:
        amts = _find_money_amounts_any(lines)
        total = max(amts) if amts else None

    if total is None or total <= 0:
        return None

    return {
        "nro": nro,
        "razon": razon,
        "cuit": cuit,
        "iibb": iibb,
        "domicilio": domicilio,
        "localidad": localidad,
        "provincia": provincia or PROVINCIA_FIJA,
        "fecha": fecha,
        "periodo": None,
        "total": float(total)
    }


def parse_formato_afip_original_block(text: str):
    if not text:
        return None

    lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
    t = _compact_ocr_token_text(" ".join(lines))

    m = re.search(
        r"Punto\s*de\s*Venta\s*:\s*(\d+)\s*(?:Comp\.?\s*)?(?:Nro|N°|Nº|No|N0)\s*:\s*(\d+)",
        t,
        re.IGNORECASE
    )
    if not m:
        return None
    nro = f"{strip_leading_zeros(m.group(1))}-{strip_leading_zeros(m.group(2))}"

    razon = ""
    for ln in lines:
        ln_norm = _compact_ocr_token_text(ln)
        rm = re.search(
            r"Raz(?:[oóó0]|6)n\s+Social\s*:\s*(.+?)\s+Fecha\s+de\s+Emisi[oóó0]n\s*:",
            ln_norm,
            re.IGNORECASE
        )
        if rm:
            razon = rm.group(1).strip()
            break

    if not razon:
        idx = None
        for i, ln in enumerate(lines):
            ln_up = _compact_ocr_token_text(ln).upper().replace(" ", "")
            if ln_up in ("ORIGINAL", "DUPLICADO", "TRIPLICADO"):
                idx = i
                break
        if idx is not None:
            for j in range(idx + 1, min(idx + 8, len(lines))):
                cand = lines[j].strip()
                if _is_bad_singleton(cand):
                    continue
                if "PUNTO DE VENTA" in cand.upper():
                    break
                razon = cand
                break

    if not razon or _is_bad_singleton(razon):
        return None

    fecha = None
    for ln in lines:
        ln_norm = _compact_ocr_token_text(ln)
        fm = re.search(r"Fecha\s+de\s+Emisi[oóó0]n\s*:\s*(\d{1,2}/\d{1,2}/\d{4})", ln_norm, re.IGNORECASE)
        if fm:
            try:
                fecha = parse_date_flexible(fm.group(1))
            except Exception:
                fecha = None
            break

    periodo = None
    for ln in lines:
        ln_norm = _compact_ocr_token_text(ln)
        pm = re.search(r"Per[ií]odo\s+Facturado\s+Desde\s*:\s*(\d{1,2}/\d{1,2}/\d{4})", ln_norm, re.IGNORECASE)
        if pm:
            try:
                periodo = parse_date_flexible(pm.group(1))
            except Exception:
                periodo = None
            break

    domicilio = ""
    localidad = ""
    provincia = PROVINCIA_FIJA
    cuit = ""
    iibb = ""

    dom_line = ""
    dom_idx = None
    for i, ln in enumerate(lines):
        if re.search(r"Domicilio\s+Comercial\s*:", ln, re.IGNORECASE):
            dom_idx = i
            dom_line = re.sub(r"^Domicilio\s+Comercial\s*:\s*", "", ln, flags=re.IGNORECASE).strip()
            break

    if dom_idx is not None and dom_idx + 1 < len(lines):
        ln2 = lines[dom_idx + 1]
        if re.search(r"^[A-Za-zÁÉÍÓÚÜÑ\s]+,\s*[A-Za-zÁÉÍÓÚÜÑ\s]+", ln2):
            locprov = re.split(r"\s+Ingresos\s+Brutos\s*:", ln2, flags=re.IGNORECASE)[0].strip()
            dom_line = (dom_line + " " + locprov).strip()

        im = re.search(r"Ingresos\s+Brutos\s*:\s*([0-9\.\-]+)", ln2, re.IGNORECASE)
        if im:
            iibb = im.group(1).strip()

    if dom_idx is not None:
        cm = re.search(r"\bCUIT\s*:\s*([0-9]{11})\b", lines[dom_idx], re.IGNORECASE)
        if cm:
            cuit = format_cuit(cm.group(1))

    if dom_line:
        domicilio, localidad, provincia = split_domicilio_localidad_provincia(dom_line)

    total = None
    for ln in lines:
        tm = re.search(r"Importe\s*Total\s*:\s*\$?\s*([0-9\.\,]+)", ln, re.IGNORECASE)
        if tm:
            total = monto_ar_to_float(tm.group(1))
            if total > 0:
                break
    if total is None or total <= 0:
        amts = _find_money_amounts_any(lines)
        total = max(amts) if amts else None

    if total is None or total <= 0:
        return None

    return {
        "nro": nro,
        "razon": razon,
        "cuit": cuit,
        "iibb": iibb,
        "domicilio": domicilio,
        "localidad": localidad,
        "provincia": provincia or PROVINCIA_FIJA,
        "fecha": fecha,
        "periodo": periodo,
        "total": float(total)
    }


def parse_formato_san_luis(text: str):
    """
    Formato tipo:
      Inicio Actividades: 06/06/2024
      CUIT: 30-71861570-0
      Fecha: 06/04/2026
      Factura
      Nro.:00102 - 00000956
      ...
      RAZON SOCIAL: ENTE DE RECUPERACION DE FONDOS ...
      ...
      Periodo de Facturacion 27-05-2025 a 24-06-2025
      Total $ 6.851.677,79
    """
    if not text:
        return None

    lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
    raw = "\n".join(lines)
    up = re.sub(r"\s+", " ", raw).upper()

    # -------- Variante histórica AFIP/San Luis (regresión) --------
    # Ejemplo: "Punto de venta ... Comp. Nro ... Razón social ... Domicilio ..."
    if re.search(r"Punto\s*de\s*venta\s*:", raw, re.IGNORECASE) and re.search(
        r"Comp\.?\s*Nro\s*:", raw, re.IGNORECASE
    ):
        m = re.search(
            r"Punto\s*de\s*venta\s*:\s*(\d+)\s+Comp\.?\s*Nro\s*:\s*(\d+)",
            raw,
            re.IGNORECASE
        )
        if m:
            nro = f"{strip_leading_zeros(m.group(1))}-{strip_leading_zeros(m.group(2))}"

            razon = ""
            domicilio_raw = ""
            rm = re.search(
                r"Raz[oó]n\s*social\s*:\s*(.*?)\s+Domicilio\s*:\s*(.*?)\s+Fecha\s*de\s*emisi[oó]n\s*:",
                raw,
                re.IGNORECASE | re.DOTALL
            )
            if rm:
                razon = re.sub(r"\s+", " ", rm.group(1).strip())
                domicilio_raw = rm.group(2).strip()

            if razon:
                domicilio = ""
                localidad = ""
                provincia = PROVINCIA_FIJA

                dom_lines = [dl.strip() for dl in domicilio_raw.splitlines() if dl.strip()]
                if dom_lines:
                    domicilio = re.sub(r"\s+", " ", dom_lines[0]).strip()
                if len(dom_lines) > 1:
                    mloc = re.search(r"^(.+?)\s*,\s*(.+)$", dom_lines[1])
                    if mloc:
                        localidad = mloc.group(1).strip()
                        provincia = mloc.group(2).strip() or PROVINCIA_FIJA
                    else:
                        localidad = re.sub(r"\s+", " ", dom_lines[1]).strip()

                if not domicilio:
                    domicilio, localidad, provincia = split_domicilio_localidad_provincia(
                        re.sub(r"\s+", " ", domicilio_raw)
                    )

                fecha = None
                fm = re.search(r"Fecha\s*de\s*emisi[oó]n\s*:\s*(\d{1,2}/\d{1,2}/\d{4})", raw, re.IGNORECASE)
                if fm:
                    try:
                        fecha = parse_date_flexible(fm.group(1))
                    except Exception:
                        fecha = None

                cuit = ""
                cm = re.search(r"\bCuit\s*:\s*([0-9]{11})\b", raw, re.IGNORECASE)
                if cm:
                    cuit = format_cuit(cm.group(1))

                iibb = ""
                im = re.search(r"Ingresos\s*brutos\s*:\s*(.+)$", raw, re.IGNORECASE | re.MULTILINE)
                if im:
                    iibb = im.group(1).strip()
                    iibb = re.split(
                        r"\s+Periodo\s+facturado|\s+Importe\s+total|\s+CAE\b",
                        iibb,
                        flags=re.IGNORECASE
                    )[0].strip()

                periodo = None
                pm = re.search(
                    r"Periodo\s+facturado\s+desde\s*:\s*(\d{1,2}/\d{1,2}/\d{4})\s+Hasta\s*:\s*(\d{1,2}/\d{1,2}/\d{4})",
                    raw,
                    re.IGNORECASE
                )
                if pm:
                    try:
                        periodo = parse_date_flexible(pm.group(1))
                    except Exception:
                        periodo = None

                total = None
                tm = re.search(r"Importe\s*total\s*:\s*([0-9\.\,]+)", raw, re.IGNORECASE)
                if tm:
                    total = monto_ar_to_float(tm.group(1))
                if total is None or total <= 0:
                    amts = _find_money_amounts_any(lines)
                    total = max(amts) if amts else None
                if total is not None and total > 0:
                    return {
                        "nro": nro,
                        "razon": razon,
                        "cuit": cuit,
                        "iibb": iibb,
                        "domicilio": domicilio,
                        "localidad": localidad,
                        "provincia": provincia or PROVINCIA_FIJA,
                        "fecha": fecha,
                        "periodo": periodo,
                        "total": float(total)
                    }

    # señales claras del formato San Luis REFORSAL
    if "RAZON SOCIAL:" not in up:
        return None
    if "NRO.:" not in up and "NRO :" not in up and "NRO:" not in up:
        return None
    if "PERIODO DE FACTURACION" not in up:
        return None
    if "TOTAL $" not in up and not re.search(r"\bTOTAL\b", up):
        return None

    # número
    m = re.search(r"Nro\.?\s*:\s*(\d+)\s*-\s*(\d+)", raw, re.IGNORECASE)
    if not m:
        return None
    nro = f"{strip_leading_zeros(m.group(1))}-{strip_leading_zeros(m.group(2))}"

    # fecha
    fecha = None
    fm = re.search(r"\bFecha\s*:\s*(\d{1,2}/\d{1,2}/\d{4})\b", raw, re.IGNORECASE)
    if fm:
        try:
            fecha = parse_date_flexible(fm.group(1))
        except Exception:
            fecha = None

    # período
    periodo = None
    pm = re.search(
        r"Periodo\s+de\s+Facturacion\s+(\d{2}-\d{2}-\d{4})\s+a\s+(\d{2}-\d{2}-\d{4})",
        raw,
        re.IGNORECASE
    )
    if pm:
        try:
            periodo = parse_date_dash(pm.group(1))
        except Exception:
            periodo = None

    # razón social
    razon = ""
    rm = re.search(
        r"RAZON\s+SOCIAL\s*:\s*(.+?)(?:\s+IVA\s*:|\nIVA\s*:|\s+REFORSAL\b|\nREFORSAL\b)",
        raw,
        re.IGNORECASE | re.DOTALL
    )
    if rm:
        razon = re.sub(r"\s+", " ", rm.group(1)).strip()
    else:
        rm2 = re.search(r"RAZON\s+SOCIAL\s*:\s*(.+)", raw, re.IGNORECASE)
        if rm2:
            razon = re.sub(r"\s+", " ", rm2.group(1)).strip()

    if not razon:
        return None

    # cuit proveedor: tomar el que aparece arriba del bloque
    cuit = ""
    cm = re.search(r"\bCUIT\s*:\s*([0-9\-]{11,14})\b", raw, re.IGNORECASE)
    if cm:
        cuit = format_cuit(cm.group(1))

    # domicilio/localidad/provincia
    domicilio = ""
    localidad = ""
    provincia = PROVINCIA_FIJA

    # línea de dirección + línea siguiente ciudad/provincia
    idx_razon = None
    for i, ln in enumerate(lines):
        if re.search(r"RAZON\s+SOCIAL\s*:", ln, re.IGNORECASE):
            idx_razon = i
            break

    if idx_razon is not None:
        for j in range(max(0, idx_razon - 3), idx_razon):
            cand = lines[j]
            if re.search(r"\d", cand) and not re.search(r"CUIT|FECHA|NRO|CODIGO|FACTURA", cand, re.IGNORECASE):
                domicilio = cand.strip()
                if j + 1 < len(lines):
                    sig = lines[j + 1].strip()
                    mloc = re.search(r"^\d{4,5}\s*-\s*(.+?)\s+(Mendoza|San Luis|CABA|Buenos Aires)$", sig, re.IGNORECASE)
                    if mloc:
                        localidad = mloc.group(1).strip()
                        provincia = mloc.group(2).strip()
                break

    # total
    total = None
    tm = re.search(r"\bTotal\s*\$\s*([0-9\.,]+)", raw, re.IGNORECASE)
    if tm:
        total = monto_to_float_any(tm.group(1))
    if total is None or total <= 0:
        amts = _find_money_amounts_any(lines)
        total = max(amts) if amts else None
    if total is None or total <= 0:
        return None

    return {
        "nro": nro,
        "razon": razon,
        "cuit": cuit,
        "iibb": "",
        "domicilio": domicilio,
        "localidad": localidad,
        "provincia": provincia or PROVINCIA_FIJA,
        "fecha": fecha,
        "periodo": periodo,
        "total": float(total)
    }


def parse_formato_hospital_ramon_carrillo(text: str):
    """
    ENTE HOSPITAL RAMON CARRILLO (formato específico).
    Detección conservadora para evitar colisiones con parsers genéricos.
    """
    if not text:
        return None

    lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
    raw = "\n".join(lines)
    up = re.sub(r"\s+", " ", raw).upper()

    # Señales específicas y combinadas del formato
    if "ENTE HOSPITAL RAMON CARRILLO" not in up:
        return None
    if not re.search(r"\bB\s*FACTURA\b", up):
        return None
    if not re.search(r"\bN[UÚ]MERO\s*:", raw, re.IGNORECASE):
        return None
    if not re.search(r"\bC\.?\s*U\.?\s*I\.?\s*T\.?\s*:", raw, re.IGNORECASE):
        return None
    if "F.VTO:" not in up:
        return None

    # Número: Numero/Número: 00003-00001755 -> 3-1755
    m = re.search(r"\bN[UÚ]MERO\s*:\s*(\d+)\s*-\s*(\d+)", raw, re.IGNORECASE)
    if not m:
        return None
    nro = f"{strip_leading_zeros(m.group(1))}-{strip_leading_zeros(m.group(2))}"

    razon = "ENTE HOSPITAL RAMON CARRILLO"

    # Fecha
    fecha = None
    fm = re.search(r"\bFecha\s*:\s*(\d{1,2}/\d{1,2}/\d{4})", raw, re.IGNORECASE)
    if fm:
        try:
            fecha = parse_date_flexible(fm.group(1))
        except Exception:
            fecha = None

    # CUIT emisor (el primero junto al bloque del emisor)
    cuit = ""
    cm = re.search(r"\bC\.?\s*U\.?\s*I\.?\s*T\.?\s*:\s*([0-9][0-9\-\.\s]{9,16}[0-9])", raw, re.IGNORECASE)
    if cm:
        cuit = format_cuit(cm.group(1))

    # IIBB
    iibb = ""
    im = re.search(r"Ing\.?\s*Brutos\s*:\s*([A-Za-z0-9\.\-]+)", raw, re.IGNORECASE)
    if im:
        iibb = im.group(1).strip()

    # Domicilio/localidad/provincia del emisor
    domicilio = ""
    localidad = ""
    provincia = PROVINCIA_FIJA

    for i, ln in enumerate(lines):
        if "ENTE HOSPITAL RAMON CARRILLO" in ln.upper():
            if i + 1 < len(lines):
                dom_ln = re.sub(r"\s+", " ", lines[i + 1]).strip()
                dm = re.match(r"(.+?)\s*-\s*([A-Za-zÁÉÍÓÚÜÑáéíóúüñ ]+)$", dom_ln)
                if dm:
                    domicilio = dm.group(1).strip(" -,")
                    localidad = dm.group(2).strip(" -,")
                else:
                    domicilio = dom_ln.strip(" -,")

            if i + 2 < len(lines):
                prov_ln = re.sub(r"\s+", " ", lines[i + 2]).strip()
                pm = re.search(r"\)\s*([A-Za-zÁÉÍÓÚÜÑáéíóúüñ ]+?)\s+Argentina\b", prov_ln, re.IGNORECASE)
                if pm:
                    provincia = pm.group(1).strip(" -,")
            break

    # Período (conservador): inferir desde Observaciones o descripción
    periodo = None
    meses = {
        "ENERO": 1, "FEBRERO": 2, "MARZO": 3, "ABRIL": 4, "MAYO": 5, "JUNIO": 6,
        "JULIO": 7, "AGOSTO": 8, "SEPTIEMBRE": 9, "SETIEMBRE": 9,
        "OCTUBRE": 10, "NOVIEMBRE": 11, "DICIEMBRE": 12
    }
    pmatch = re.search(
        r"\b(" + "|".join(meses.keys()) + r")\s+(20\d{2})\b",
        raw.upper(),
        re.IGNORECASE
    )
    if pmatch:
        mm = meses.get(pmatch.group(1).upper())
        yy = int(pmatch.group(2))
        if mm:
            periodo = datetime(yy, mm, 1)

    # Total
    total = None
    tm = re.search(r"\bTotal\s*:\s*\$?\s*([0-9\.,]+)", raw, re.IGNORECASE)
    if tm:
        total = monto_to_float_any(tm.group(1))
    if total is None or total <= 0:
        amts = _find_money_amounts_any(lines)
        total = max(amts) if amts else None
    if total is None or total <= 0:
        return None

    return {
        "nro": nro,
        "razon": razon,
        "cuit": cuit,
        "iibb": iibb,
        "domicilio": domicilio,
        "localidad": localidad,
        "provincia": provincia or PROVINCIA_FIJA,
        "fecha": fecha,
        "periodo": periodo,
        "total": float(total)
    }


def parse_formato_afip(text: str):
    if not text:
        return None

    lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
    t = " ".join(lines)

    m = re.search(
        r"Punto\s*de\s*Venta\s*:\s*(\d+)\s*"
        r"(?:Comp\.?\s*(?:Nro|Nº|N°|No)|Comp\.)\s*[:.]?\s*(\d+)",
        t, re.IGNORECASE
    )
    if not m:
        return None

    nro = f"{strip_leading_zeros(m.group(1))}-{strip_leading_zeros(m.group(2))}"

    proveedor = ""
    idx_pv = None
    for i, ln in enumerate(lines):
        if re.search(r"Punto\s*de\s*Venta\s*:", ln, re.IGNORECASE):
            idx_pv = i
            break
    if idx_pv is not None and idx_pv + 1 < len(lines):
        proveedor = lines[idx_pv + 1].strip()
    if ":" in (proveedor or ""):
        proveedor = ""

    fecha = None
    for ln in lines:
        fm = re.search(r"Fecha\s*(?:de\s*)?Emisi\w*\s*:\s*(\d{1,2}/\d{1,2}/\d{4})", ln, re.IGNORECASE)
        if fm:
            fecha = parse_date_flexible(fm.group(1))
            break

    domicilio = ""
    cuit = ""
    for ln in lines:
        if re.search(r"^Domicilio\s*:", ln, re.IGNORECASE) and re.search(r"\bCUIT\s*:", ln, re.IGNORECASE):
            dm = re.search(r"Domicilio\s*:\s*(.+?)\s*CUIT\s*:\s*([0-9\-]{11,13})", ln, re.IGNORECASE)
            if dm:
                domicilio = dm.group(1).strip()
                cuit = format_cuit(dm.group(2))
                break

    iibb = ""
    for ln in lines:
        im = re.search(r"Ingresos\s*Brutos\s*:\s*([0-9\-\.]+)", ln, re.IGNORECASE)
        if im:
            iibb = im.group(1).strip()
            break

    tm = re.search(r"Importe\s*Total\s*:\s*\$?\s*([0-9\.\,]+)", t, re.IGNORECASE)
    if not tm:
        return None
    total = monto_ar_to_float(tm.group(1))
    if total <= 0:
        return None
    if not proveedor:
        return None

    return {
        "nro": nro,
        "razon": proveedor,
        "cuit": cuit,
        "iibb": iibb,
        "domicilio": domicilio,
        "localidad": "",
        "provincia": PROVINCIA_FIJA,
        "fecha": fecha,
        "periodo": None,
        "total": float(total)
    }


def parse_formato_nuevo(text: str):
    if not text:
        return None

    lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
    raw = "\n".join(lines)
    t = re.sub(r"\s+", " ", raw)

    m = re.search(r"\bNro\.\s*:\s*(\d+)\s*-\s*(\d+)\b", t, re.IGNORECASE)
    if not m:
        return None
    nro = f"{strip_leading_zeros(m.group(1))}-{strip_leading_zeros(m.group(2))}"

    fecha_m = re.search(r"\bFecha:\s*(\d{1,2}/\d{1,2}/\d{4})\b", t, re.IGNORECASE)
    fecha = parse_date_flexible(fecha_m.group(1)) if fecha_m else None

    total_m = re.search(r"\bTotal\b\s*\$?\s*([0-9\.\,]+)", t, re.IGNORECASE)
    if not total_m:
        return None
    total = monto_ar_to_float(total_m.group(1))
    if total <= 0:
        return None

    per = re.search(r"Periodo de Facturacion\s+(\d{2}-\d{2}-\d{4})\s+a\s+(\d{2}-\d{2}-\d{4})", t, re.IGNORECASE)
    periodo = parse_date_dash(per.group(1)) if per else None

    bloque = ""
    bm = re.search(r"RAZON SOCIAL:\s*(.+?)(?:Inicio Actividades:|IVA:)", raw, re.IGNORECASE | re.DOTALL)
    if bm:
        bloque = re.sub(r"\s+", " ", bm.group(1)).strip()
    else:
        bm2 = re.search(r"RAZON SOCIAL:\s*(.+?)(?:\s+IVA:| IVA:)", raw, re.IGNORECASE | re.DOTALL)
        bloque = re.sub(r"\s+", " ", bm2.group(1)).strip() if bm2 else ""

    proveedor = bloque
    if "CUIT:" in bloque.upper():
        proveedor = re.split(r"CUIT:", bloque, flags=re.IGNORECASE)[0].strip()

    cuit = ""
    cm = re.search(r"\bCUIT:\s*([0-9\-]{11,13})", bloque, re.IGNORECASE)
    if cm:
        cuit = format_cuit(cm.group(1))

    iibb = ""
    iibbm = re.search(r"Ingresos Brutos:\s*([0-9\-\.]+)", t, re.IGNORECASE)
    if iibbm:
        iibb = iibbm.group(1).strip()

    if not proveedor:
        return None

    return {
        "nro": nro,
        "razon": proveedor,
        "cuit": cuit,
        "iibb": iibb,
        "domicilio": "",
        "localidad": "",
        "provincia": PROVINCIA_FIJA,
        "fecha": fecha,
        "periodo": periodo,
        "total": float(total)
    }


def extraer_campos(pdf_path: str, error_dir: Path, pdf_obj: Path, debug: bool = False):
    text = read_pdf_text_first_page(pdf_path)

    # ✅ NUEVO: A.M.A
    campos = parse_formato_ama(text)
    if campos:
        if debug:
            print("PARSER -> AMA")
            save_parse_debug(error_dir, pdf_obj, "AMA", campos, text)
        return campos

    # ✅ MK & G3
    campos = parse_formato_mk_g3(text)
    if campos:
        if debug:
            print("PARSER -> MK_G3")
            save_parse_debug(error_dir, pdf_obj, "MK_G3", campos, text)
        return campos

    # ✅ MENDOSALUD (específico, antes de genéricos)
    campos = parse_formato_mendosalud(text)
    if campos:
        if debug:
            print("PARSER -> MENDOSALUD")
            save_parse_debug(error_dir, pdf_obj, "MENDOSALUD", campos, text)
        return campos

    # ✅ Hospital Universitario
    campos = parse_formato_hospital_universitario(text)
    if campos:
        if debug:
            print("PARSER -> HOSPITAL_UNIVERSITARIO")
            save_parse_debug(error_dir, pdf_obj, "HOSPITAL_UNIVERSITARIO", campos, text)
        return campos

    # ✅ Nutrihome
    campos = parse_formato_nutrihome(text)
    if campos:
        if debug:
            print("PARSER -> NUTRIHOME")
            save_parse_debug(error_dir, pdf_obj, "NUTRIHOME", campos, text)
        return campos

    campos = parse_formato_afip_original_block(text)
    if campos:
        if debug:
            print("PARSER -> AFIP_ORIGINAL_BLOCK")
            save_parse_debug(error_dir, pdf_obj, "AFIP_ORIGINAL_BLOCK", campos, text)
        return campos

    campos = parse_formato_san_luis(text)
    if campos:
        if debug:
            print("PARSER -> SAN_LUIS")
            save_parse_debug(error_dir, pdf_obj, "SAN_LUIS", campos, text)
        return campos

    # ✅ Hospital Ramón Carrillo (específico, antes de genéricos)
    campos = parse_formato_hospital_ramon_carrillo(text)
    if campos:
        if debug:
            print("PARSER -> HOSPITAL_RAMON_CARRILLO")
            save_parse_debug(error_dir, pdf_obj, "HOSPITAL_RAMON_CARRILLO", campos, text)
        return campos

    campos = parse_formato_afip(text)
    if campos:
        if debug:
            print("PARSER -> AFIP")
            save_parse_debug(error_dir, pdf_obj, "AFIP", campos, text)
        return campos

    campos = parse_formato_nuevo(text)
    if campos:
        if debug:
            print("PARSER -> NUEVO")
            save_parse_debug(error_dir, pdf_obj, "NUEVO", campos, text)
        return campos

    save_debug_text(error_dir, pdf_obj, text)
    raise ValueError("Formato de factura no reconocido.")


# ================= CARPETA SALIDA =================
def carpeta_mes_anterior_por_creacion(pdf_path: str) -> Path:
    ctime = os.path.getctime(pdf_path)
    dt = datetime.fromtimestamp(ctime)

    primero_mes = dt.replace(day=1)
    dt_obj = primero_mes - timedelta(days=1)

    anio = dt_obj.year
    mes = dt_obj.month

    desktop = Path.home() / "Desktop"
    carpeta = desktop / "excels-proveedores" / str(anio) / f"{anio}-{mes:02d}_{MESES_ES[mes]}"
    carpeta.mkdir(parents=True, exist_ok=True)
    return carpeta


# ================= MAIN (AUTO-DEBUG) =================
def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--pdf", required=True)
    parser.add_argument("--plantilla", required=True)
    parser.add_argument("--error_dir", default="ERROR")
    parser.add_argument("--debug", action="store_true")
    args = parser.parse_args()

    pdf_path = Path(args.pdf)
    error_dir = Path(args.error_dir)
    error_dir.mkdir(parents=True, exist_ok=True)

    def procesar(debug_flag: bool):
        campos = extraer_campos(str(pdf_path), error_dir, pdf_path, debug=debug_flag)

        ok, msg = validar_campos_criticos(campos)
        if not ok:
            raise ValueError(f"CAMPOS CRÍTICOS FALTANTES: {msg}")

        wb = openpyxl.load_workbook(args.plantilla)
        ws = wb.active

        set_proveedor_en_A8(ws, campos["razon"])

        ws["A11"] = f"Cuit: {campos['cuit']}"
        ws["A12"] = f"Ingresos brutos: {campos['iibb']}"
        ws["A13"] = f"Domicilio: {campos['domicilio']}"
        ws["A14"] = f"Localidad: {campos['localidad']}"
        ws["A15"] = f"Provincia: {campos['provincia']}"

        fila = None
        for r in range(FILAS_INICIO, FILAS_FIN + 1):
            if not ws[f"C{r}"].value:
                fila = r
                break
        if fila is None:
            raise RuntimeError("No hay filas libres entre 9 y 21.")

        base_mes = month_start(campos["periodo"] or campos["fecha"])
        realiz = prev_month_start(base_mes) if base_mes else None

        ws[f"B{fila}"] = realiz
        ws[f"C{fila}"] = campos["nro"]
        ws[f"D{fila}"] = campos["total"]
        ws[f"E{fila}"] = 0.0
        ws[f"F{fila}"] = campos["fecha"]

        if realiz:
            ws[f"B{fila}"].number_format = "dd/mm/yyyy"
        ws[f"D{fila}"].number_format = "#,##0.00"
        ws[f"E{fila}"].number_format = "#,##0.00"
        if campos["fecha"]:
            ws[f"F{fila}"].number_format = "dd/mm/yyyy"

        ws["A40"] = total_pesos_a_letras(campos["total"])

        outdir = carpeta_mes_anterior_por_creacion(str(pdf_path))
        proveedor_arch = sanitize_filename(campos["razon"])
        excel_name = f"{proveedor_arch} - {campos['nro']}.xlsx"
        out_path = unique_path(outdir / excel_name)

        wb.save(out_path)
        print("OK EXCEL ->", out_path)

        if IMPRIMIR_AUTOMATICO:
            ok_print = imprimir_excel_windows(str(out_path), copies=COPIAS_IMPRESION)
            if ok_print:
                print(f"🖨 Enviado a imprimir ({COPIAS_IMPRESION} copia/s)")
            else:
                print("⚠ No pude imprimir automáticamente (falló COM y startfile).")

        try:
            pdf_path.unlink()
        except Exception:
            dest = unique_path(error_dir / pdf_path.name)
            try:
                shutil.move(str(pdf_path), str(dest))
                print("AVISO: No pude borrar el PDF (en uso). Lo moví a ERROR ->", dest)
            except Exception:
                print("AVISO: No pude borrar ni mover el PDF. Puede estar en uso.")

    try:
        procesar(debug_flag=args.debug)

    except Exception as e1:
        print("ERROR:", e1)

        if not args.debug:
            try:
                print("🔎 Auto-debug activado (reintento con --debug para diagnóstico)...")
                procesar(debug_flag=True)
                return
            except Exception as e2:
                print("ERROR (auto-debug):", e2)

        if pdf_path.exists():
            try:
                dest = unique_path(error_dir / pdf_path.name)
                shutil.move(str(pdf_path), str(dest))
                print("PDF -> ERROR:", dest)
            except Exception:
                pass
        raise


if __name__ == "__main__":
    main()
